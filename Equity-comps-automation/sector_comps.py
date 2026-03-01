import yfinance as yf
import pandas as pd


def get_sector_comps(ticker_list):

    # ── Fix #6: Input Validation ──────────────────────────────────────────────
    # Without this, an empty list or junk input silently produces a broken
    # DataFrame with NaN medians, which then crashes when building the median row.
    if not isinstance(ticker_list, (list, tuple)):
        raise TypeError(
            f"ticker_list must be a list or tuple, got {type(ticker_list).__name__}."
        )

    ticker_list = [t for t in ticker_list if isinstance(t, str) and t.strip()]

    if not ticker_list:
        raise ValueError(
            "ticker_list is empty or contains no valid ticker strings. "
            "Please provide at least one ticker symbol e.g. ['CRM', 'NOW']."
        )
    # ─────────────────────────────────────────────────────────────────────────

    data_list = []

    for ticker in ticker_list:
        try:
            company = yf.Ticker(ticker)
            info = company.get_info()

            market_cap = info.get("marketCap")
            ev = info.get("enterpriseValue")
            ebitda = info.get("ebitda")
            revenue_growth = info.get("revenueGrowth")
            pe_ratio = info.get("trailingPE")

            market_cap_formatted = f"${round(market_cap/1e9,1)}B" if market_cap else None
            ev_ebitda = round(ev/ebitda,2) if ev and ebitda else None
            revenue_growth = round(revenue_growth*100,2) if revenue_growth else None
            pe_ratio = round(pe_ratio,2) if pe_ratio else None

            data_list.append({
                "Ticker": ticker,
                "Company": info.get("longName"),
                "Market Cap ($B)": market_cap_formatted,
                "P/E": pe_ratio,
                "Revenue Growth (%)": revenue_growth,
                "EV/EBITDA": ev_ebitda
            })

        except Exception as e:
            print(f"Error pulling {ticker}: {e}")

    df = pd.DataFrame(data_list)

    # Sector Medians
    sector_median_pe = df["P/E"].median()
    sector_median_ev_ebitda = df["EV/EBITDA"].median()
    sector_median_growth = df["Revenue Growth (%)"].median()

    # A low P/E can equally signal
    # slow growth, sector mismatch, or financial distress. "Low P/E vs. Peers"
    # is factually accurate and defensible; it describes what the data shows
    # without implying a buy recommendation.
    df["Screening Flag"] = df["P/E"].apply(
        lambda x: "Low P/E vs. Peers" if x and x < 0.8 * sector_median_pe else "—"
    )
    # ─────────────────────────────────────────────────────────────────────────

    # Add Median Row
    median_row = {
        "Ticker": "",
        "Company": "SECTOR MEDIAN",
        "Market Cap ($B)": "",
        "P/E": round(sector_median_pe, 2),
        "Revenue Growth (%)": round(sector_median_growth, 2),
        "EV/EBITDA": round(sector_median_ev_ebitda, 2),
        "Screening Flag": ""
    }

    df = pd.concat([df, pd.DataFrame([median_row])], ignore_index=True)

    return df


if __name__ == "__main__":
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    saas_tickers = ["NVDA", "AAPL", "MSFT", "AMZN", "GOOGL", "META", "TSLA", "AVGO", "NFLX", "AMD", "INTC", "QCOM", "ADBE", "CRM", "ORCL", "CSCO", "TXN", "MU", "LRCX", "AMAT"]

    # ─────────────────────────────────────────────────────────────────────────

    file_name = "SaaS_Comps_Function_Version.xlsx"

    df = get_sector_comps(saas_tickers)

    # Start dataset at B2
    df.to_excel(file_name, index=False, startrow=1, startcol=1)

    wb = load_workbook(file_name)
    ws = wb.active

    # Remove gridlines
    ws.sheet_view.showGridLines = False

    # Set column A narrow spacer
    ws.column_dimensions["A"].width = 2

    # Freeze panes (B3 because data starts at B2)
    ws.freeze_panes = "B3"

    # Header styling
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    for cell in ws[2]:  # Header is now row 2
        if cell.column >= 2:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    # Thin institutional borders
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=3):
        for cell in row:
            if cell.column >= 2:
                cell.border = border

                # Right align numeric columns
                if cell.column > 4:
                    cell.alignment = Alignment(horizontal="right")

                    # Apply finance-style number format
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00;(#,##0.00)'
                else:
                    cell.alignment = Alignment(horizontal="left")

    # Auto-adjust widths (excluding spacer column A)
    for col in ws.columns:
        if col[0].column >= 2:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = max_length + 3

    # Highlight "Low P/E vs. Peers" rows (previously "Undervalued")
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    for row in ws.iter_rows(min_row=3):
        if row[-1].value == "Low P/E vs. Peers":
            row[-1].fill = green_fill
            row[-1].font = Font(bold=True)

    # Format Sector Median row
    grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for row in ws.iter_rows():
        if row[2].value == "SECTOR MEDIAN":
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = grey_fill

    wb.save(file_name)

    print("Institutional formatted Excel successfully created.")